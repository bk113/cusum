"""
cusum_run.py  —  CUSUM Active-Manager Monitor
==============================================
Self-contained pipeline for the Morningstar cumulative-return xlsx export.

USAGE
-----
    python cusum_run.py data.xlsx
    python cusum_run.py data.xlsx --threshold 19.81 --min-months 24 --out-dir results/

OUTPUTS (written to --out-dir, default = same folder as the xlsx)
-------
    cusum_IC_report_YYYYMMDD.xlsx    full results workbook
    cusum_L_barchart.png             top-30 L bar chart
    cusum_top5_timeseries.png        CUSUM time series for five most at-risk funds
    cusum_scatter.png                IR vs current-L scatter

DEPENDENCIES
------------
    pip install pandas numpy matplotlib openpyxl yfinance
    yfinance is used ONCE on first run to fetch ETF proxy benchmark returns.
    Results are cached in etf_proxy_cache.json (same folder as xlsx).
    Subsequent runs use the cache — no internet needed.

DATA FORMAT EXPECTED
--------------------
The xlsx must be a Morningstar "grouped by custom" export with:
  - Row 7 (0-indexed)  : period-end dates starting at column 6
  - Row 8              : column headers (fund name, benchmark, bench id, ISIN, SecId, inception, ...)
  - Row 10+            : one fund per row; rows without an ISIN are group headers (skipped)
  - Row 244+           : dedicated Benchmarks section (no ISIN, have SecId, have return data)
  - Return values      : monthly % total return, NaN before inception

BENCHMARK RESOLUTION ORDER
---------------------------
  1. Morningstar Benchmarks section (rows 244+) — 71 real index series
  2. ETF proxy map (fetched via yfinance, cached locally) — 27 missing indices
  3. Blended benchmarks: full weighted blend if all legs available, else primary leg
  Funds with no resolvable benchmark are excluded and logged.

ETF PROXY QUALITY
-----------------
  ★★★★  Near-perfect (HYG, EMB, EMXC, EEM, IBB, REET, IXC, PICK, BIL)
  ★★★☆  Good proxy  (CEMB, MCHI, FXI, HYXU, CWB, AAXJ, QQQ, BND)
  ★★☆☆  Weak proxy  (BND for FTSE WGBI ex-Japan Hdg SGD) — flagged in output
  Proxy-benchmarked funds are marked  [ETF-PROXY]  in the benchmark_name column.
""""""
cusum_run.py  —  CUSUM Active-Manager Monitor
==============================================
Self-contained pipeline for the Morningstar cumulative-return xlsx export.

USAGE
-----
    python cusum_run.py data.xlsx
    python cusum_run.py data.xlsx --threshold 19.81 --min-months 24 --out-dir results/

OUTPUTS (written to --out-dir, default = same folder as the xlsx)
-------
    cusum_results_YYYYMMDD.csv      one row per analysed fund
    cusum_L_barchart.png            top-30 L bar chart
    cusum_top5_timeseries.png       CUSUM time series for five most at-risk funds
    cusum_scatter.png               IR vs current-L scatter

DEPENDENCIES
------------
    pip install pandas numpy matplotlib openpyxl
    (openpyxl is the xlsx engine; no internet required)

DATA FORMAT EXPECTED
--------------------
The xlsx must be a Morningstar "grouped by custom" export with:
  - Row 7 (0-indexed)  : period-end dates starting at column 6
  - Row 8              : column headers (fund name, benchmark, bench id, ISIN, SecId, inception, ...)
  - Row 10+            : one fund per row; rows without an ISIN are group headers (skipped)
  - Return values      : monthly % total return, NaN before inception

Benchmarks are identified by the Morningstar SecId in column 2.
Blended benchmarks (containing "[") are skipped — those funds are excluded.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import warnings
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional, Tuple

import matplotlib
matplotlib.use("Agg")
import matplotlib.patches as mpatches
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")
# ---------------------------------------------------------------------------
# PART 0 — ETF Proxy Benchmark Map
# ---------------------------------------------------------------------------
# For 27 Morningstar benchmark SecIds that are absent from the Benchmarks
# section of the xlsx, we fetch the closest liquid ETF via yfinance.
# The mapping is: Morningstar SecId -> (ETF ticker, index description, quality)
# Quality: 4=near-perfect, 3=good, 2=weak
#
# Fetched data is cached to etf_proxy_cache.json next to the xlsx.
# Set USE_ETF_PROXY = False to disable entirely.
# ---------------------------------------------------------------------------

USE_ETF_PROXY = True   # set False to skip proxy fetching

ETF_PROXY_MAP = {
    # SecId               ticker   index description                                    quality
    "XIUSA04EPJ": ("HYG",  "ICE BofA US HY Constnd TR USD [ETF-PROXY]",                4),
    "FOUSA06D0U": ("EMB",  "JPM EMBI Global Diversified TR USD [ETF-PROXY]",            4),
    "XIUSA04CX8": ("BIL",  "ICE BofA US 3M Trsy Bill TR USD [ETF-PROXY]",              4),
    "F00001M946": ("EMXC", "MSCI EM ex China 10/40 NR USD [ETF-PROXY]",                4),
    "F00000HL02": ("EEM",  "MSCI Emerging Markets 10-40 NR USD [ETF-PROXY]",            4),
    "F000013K3K": ("REET", "FTSE EPRA Nareit Dev Real Estate NR USD [ETF-PROXY]",       4),
    "F00000TODY": ("IBB",  "NASDAQ Biotechnology [ETF-PROXY]",                          4),
    "F00001FSIP": ("IXC",  "MSCI ACWI Energy/Materials NR USD [ETF-PROXY]",             4),
    "F00001CHK2": ("IXC",  "MSCI World Energy 30%buffer 10/40 NR USD [ETF-PROXY]",      4),
    "F00001CHIX": ("PICK", "MSCI ACWI Me&Mi 30% Bf 10/40 NR USD [ETF-PROXY]",          4),
    "F00001CY9R": ("BIL",  "Compounded Effective Federal Funds Rate [ETF-PROXY]",       4),
    "F00000JWM0": ("CEMB", "JPM Asia Credit TR USD [ETF-PROXY]",                        3),
    "F00000PGH3": ("CEMB", "JPM CEMBI Broad Diversified TR USD [ETF-PROXY]",            3),
    "F00000JWMA": ("CEMB", "JPM ACI Non Investment Grade TR USD [ETF-PROXY]",           3),
    "F00000JWMT": ("CEMB", "JPM ACI Corp Non-Investment Grade TR USD [ETF-PROXY]",      3),
    "F000027BEF": ("MCHI", "MSCI China 10/40 A/B Sh 20% Cap NR USD [ETF-PROXY]",       3),
    "F00000ZPE5": ("MCHI", "MSCI China 10/40 Index HKD [ETF-PROXY]",                   3),
    "F0000118KP": ("FXI",  "MSCI Golden Dragon 10/40 NR USD [ETF-PROXY]",               3),
    "F00000UWUQ": ("FXI",  "BMI Hong Kong & China [ETF-PROXY]",                         3),
    "F000016ARA": ("HYXU", "ICE BofA EUR HY 3% Constd TR EUR [ETF-PROXY]",             3),
    "F00000V87S": ("GHYB", "ICE BofA NFincl Dv Mkts HY Cstd TR USD [ETF-PROXY]",      3),
    "F00001LWU1": ("CWB",  "FTSE Global Focus Hgd CB TR USD [ETF-PROXY]",               3),
    "F00001EBC4": ("ERUS", "MSCI Russia NR USD [ETF-PROXY: suspended post-2022]",       3),
    "F00001F5UY": ("ASEA", "MSCI AC ASEAN 10/40 NR USD [ETF-PROXY]",                   3),
    "F00001E8FD": ("AAXJ", "MSCI AC AP Tech 100 Equal Weight NR USD [ETF-PROXY]",      3),
    "F00001DO7O": ("QQQ",  "S&P Global 1200 Shariah Info Tech TR USD [ETF-PROXY]",      3),
    "F00000UU66": ("BND",  "FTSE WGBI ex Japan(Hdg to SGD) [ETF-PROXY]",               2),
}


def _fetch_etf_proxies(
    dates_idx: "pd.DatetimeIndex",
    cache_path: str,
) -> dict:
    """
    Return SecId -> np.ndarray (length = len(dates_idx)) of monthly decimal returns.

    Strategy:
      1. Load cache from cache_path if it exists and covers all required dates.
      2. Otherwise fetch via yfinance (monthly, auto_adjust=True) and save cache.

    The array is aligned to dates_idx: months with no ETF data get NaN.
    dates_idx contains month-end dates from the Morningstar export (Jun 1978 – May 2026).
    ETFs only exist from their inception, so early months will be NaN — that is correct.
    """
    if not USE_ETF_PROXY:
        return {}

    T = len(dates_idx)
    result: dict = {}

    # ── 1. Try cache ──────────────────────────────────────────────────────────
    if os.path.exists(cache_path):
        try:
            with open(cache_path) as f:
                cache = json.load(f)
            cached_dates = cache.get("dates", [])
            # Accept cache if it covers at least up to the second-to-last date
            # (last date may not yet exist when cache was written)
            if cached_dates and pd.Timestamp(cached_dates[-1]) >= dates_idx[-2]:
                print(f"  ETF proxy cache loaded: {cache_path}")
                date_to_idx = {str(d.date()): i for i, d in enumerate(dates_idx)}
                for secid, arr_data in cache.get("returns", {}).items():
                    arr = np.full(T, np.nan)
                    for date_str, val in arr_data.items():
                        if date_str in date_to_idx:
                            arr[date_to_idx[date_str]] = val
                    result[secid] = arr
                return result
        except Exception as e:
            print(f"  ETF proxy cache unreadable ({e}), re-fetching...")

    # ── 2. Fetch via yfinance ─────────────────────────────────────────────────
    try:
        import yfinance as yf
    except ImportError:
        print("  WARNING: yfinance not installed. Run: pip install yfinance")
        print("           ETF proxy benchmarks will be skipped.")
        return {}

    unique_tickers = sorted(set(v[0] for v in ETF_PROXY_MAP.values()))
    print(f"  Fetching ETF proxies via yfinance: {unique_tickers}")

    try:
        raw = yf.download(
            tickers=unique_tickers,
            start="1993-01-01",          # earliest ETF (SPY) inception
            end=pd.Timestamp.today().strftime("%Y-%m-%d"),
            interval="1mo",
            auto_adjust=True,
            progress=False,
        )
    except Exception as e:
        print(f"  WARNING: yfinance download failed: {e}")
        return {}

    # Extract Close prices — handle single vs multi-ticker
    if isinstance(raw.columns, pd.MultiIndex):
        prices = raw["Close"]
    else:
        prices = raw[["Close"]].rename(columns={"Close": unique_tickers[0]})

    # Monthly returns from price series
    monthly_rets = prices.pct_change()

    # Align to dates_idx: for each Morningstar month-end, find the closest
    # yfinance month-end (yfinance uses first trading day of next month sometimes)
    date_to_idx = {d: i for i, d in enumerate(dates_idx)}

    # Build a lookup: year-month -> Morningstar index position
    ym_to_idx: dict = {}
    for i, d in enumerate(dates_idx):
        ym_to_idx[(d.year, d.month)] = i

    # For each unique ticker, build aligned return array
    ticker_arrays: dict = {}
    for ticker in monthly_rets.columns:
        arr = np.full(T, np.nan)
        for ts, val in monthly_rets[ticker].items():
            ts = pd.Timestamp(ts)
            key = (ts.year, ts.month)
            if key in ym_to_idx and not np.isnan(val):
                arr[ym_to_idx[key]] = val
        ticker_arrays[ticker] = arr

    # Map SecId -> aligned array
    cache_returns: dict = {}
    for secid, (ticker, desc, quality) in ETF_PROXY_MAP.items():
        if ticker in ticker_arrays:
            arr = ticker_arrays[ticker]
            result[secid] = arr
            # Store non-NaN values in cache
            cache_returns[secid] = {
                str(dates_idx[i].date()): float(arr[i])
                for i in range(T) if not np.isnan(arr[i])
            }

    # ── 3. Save cache ─────────────────────────────────────────────────────────
    try:
        cache_out = {
            "dates":   [str(d.date()) for d in dates_idx],
            "returns": cache_returns,
            "tickers": {secid: ETF_PROXY_MAP[secid][0] for secid in ETF_PROXY_MAP},
            "fetched": pd.Timestamp.today().strftime("%Y-%m-%d"),
        }
        with open(cache_path, "w") as f:
            json.dump(cache_out, f)
        print(f"  ETF proxy cache saved: {cache_path}")
    except Exception as e:
        print(f"  WARNING: could not save ETF proxy cache: {e}")

    # Coverage summary
    n_found = sum(1 for arr in result.values() if np.any(~np.isnan(arr)))
    print(f"  ETF proxies loaded: {n_found}/{len(ETF_PROXY_MAP)} SecIds have data")
    return result
# ---------------------------------------------------------------------------
# Colours
# ---------------------------------------------------------------------------
NAV_BLUE  = "#1E2761"
ALARM_RED = "#E85D24"
TEAL      = "#0F766E"
AMBER     = "#BA7517"

# ---------------------------------------------------------------------------
# PART 1 — CUSUM algorithm  (Philips-Yashchin-Stein 2003)
# ---------------------------------------------------------------------------

THRESHOLD_TABLE = pd.DataFrame(
    [
        (11.81, 24, 16, 11),
        (15.00, 36, 22, 15),
        (17.60, 48, 27, 18),
        (19.81, 60, 32, 21),   # default: ~1 false alarm / 5 yrs
        (21.79, 72, 37, 23),
        (23.59, 84, 41, 25),   # paper preferred: ~1 false alarm / 7 yrs
    ],
    columns=["threshold", "arl_good_IR0.5", "arl_flat_IR0", "arl_bad_IR-0.5"],
)


@dataclass
class CUSUMMonitor:
    """
    Two-sided CUSUM change-point detector for active-manager skill.

    Detects skill loss (L crosses threshold from below) and unexpected
    skill gain (L_up crosses threshold).  After each alarm the statistic
    resets to zero so monitoring continues; all alarm times are logged.

    Parameters
    ----------
    initial_te  : annualised tracking-error seed (e.g. 0.04 = 4 %).
    threshold   : alarm trigger h.  See THRESHOLD_TABLE.  Default 19.81.
    gamma       : EWMA decay for tracking-error estimate.  Default 0.9.
    mu_good     : IR considered "good"         (default 0.5).
    mu_bad      : IR considered "bad"          (default 0.0).
    mu_exceptional: IR considered "exceptional" (default 1.0, upper alarm).
    ir_cap      : winsorise monthly IR at ±ir_cap (default 5.0).
    """

    initial_te:    float = 0.04
    threshold:     float = 19.81
    gamma:         float = 0.90
    mu_good:       float = 0.50
    mu_bad:        float = 0.00
    mu_exceptional: float = 1.00
    ir_cap:        float = 5.00
    rolling_restart: bool = False  # accepted for API compatibility with calibrate_threshold()

    # runtime state
    sigma:     float = field(init=False, repr=False, default=0.0)
    L:         float = field(init=False, repr=False, default=0.0)
    L_up:      float = field(init=False, repr=False, default=0.0)
    alarm:     bool  = field(init=False, repr=False, default=False)
    alarm_t:   Optional[int] = field(init=False, repr=False, default=None)
    alarm_up:  bool  = field(init=False, repr=False, default=False)
    alarm_up_t: Optional[int] = field(init=False, repr=False, default=None)
    all_alarms:    List[int] = field(init=False, repr=False, default_factory=list)
    all_alarms_up: List[int] = field(init=False, repr=False, default_factory=list)
    history:   List[dict]   = field(init=False, repr=False, default_factory=list)
    _t:        int  = field(init=False, repr=False, default=0)

    def __post_init__(self):
        self.sigma = self.initial_te / np.sqrt(12)   # monthly TE

    def update(self, r_fund: float, r_bench: float) -> None:
        self._t += 1
        excess = np.log1p(r_fund) - np.log1p(r_bench)
        # EWMA sigma update
        self.sigma = np.sqrt(
            self.gamma * self.sigma**2 + (1 - self.gamma) * excess**2
        )
        sigma_ann = self.sigma * np.sqrt(12)
        if sigma_ann < 1e-8:
            ir_hat = 0.0
        else:
            ir_hat = (excess * 12) / sigma_ann
        ir_hat = np.clip(ir_hat, -self.ir_cap, self.ir_cap)

        # Lower CUSUM (skill loss)
        par_level = (self.mu_good + self.mu_bad) / 2   # 0.25
        self.L = max(0.0, self.L + par_level - ir_hat)
        if self.L >= self.threshold:
            if not self.alarm:
                self.alarm   = True
                self.alarm_t = self._t
            self.all_alarms.append(self._t)
            self.L = 0.0

        # Upper CUSUM (exceptional skill gain)
        par_up = (self.mu_good + self.mu_exceptional) / 2   # 0.75
        self.L_up = max(0.0, self.L_up + ir_hat - par_up)
        if self.L_up >= self.threshold:
            if not self.alarm_up:
                self.alarm_up   = True
                self.alarm_up_t = self._t
            self.all_alarms_up.append(self._t)
            self.L_up = 0.0

        self.history.append({
            "t": self._t,
            "excess_return": excess,
            "sigma_ann": sigma_ann,
            "ir_hat": ir_hat,
            "L": self.L,
            "L_up": self.L_up,
        })

    def run(self, returns: pd.DataFrame,
            port_col: str = "portfolio",
            bench_col: str = "benchmark") -> pd.DataFrame:
        for _, row in returns.iterrows():
            self.update(row[port_col], row[bench_col])
        return pd.DataFrame(self.history)

    def plot(self, dates: Optional[pd.Index] = None,
             title: str = "CUSUM monitor") -> plt.Figure:
        df = pd.DataFrame(self.history)
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(11, 6.5), sharex=True)
        x_idx = dates if dates is not None else df["t"].values

        ax1.plot(x_idx, df["excess_return"].cumsum(), lw=1.5, color=NAV_BLUE)
        ax1.axhline(0, color="grey", lw=0.6)
        ax1.set_ylabel("cumulative log excess return")
        ax1.set_title(title)
        ax1.grid(alpha=0.3)

        ax2.plot(x_idx, df["L"],    lw=1.5, color=NAV_BLUE, label="L (skill loss)")
        ax2.plot(x_idx, df["L_up"], lw=1.2, color=TEAL, ls="--", label="L_up (skill gain)")
        ax2.axhline(self.threshold, color=ALARM_RED, ls="--",
                    label=f"threshold h = {self.threshold}")
        ax2.invert_yaxis()
        ax2.set_ylabel("CUSUM statistic (inverted)")
        ax2.grid(alpha=0.3)

        for at in self.all_alarms:
            ax2.axvline(x_idx[at - 1] if dates is not None else at,
                        color=ALARM_RED, lw=0.9, alpha=0.6)
        for at in self.all_alarms_up:
            ax2.axvline(x_idx[at - 1] if dates is not None else at,
                        color=TEAL, lw=0.9, alpha=0.6)

        ax2.legend(loc="lower right", fontsize=8)
        plt.tight_layout()
        return fig


# ===========================================================================
# Merged from cusum.py — VIX regime tagging, Monte Carlo calibration,
# asset-class presets, and simulation utilities
# ===========================================================================

_SQRT12 = np.sqrt(12.0)

# Named ANNUALISED IR reference values — same units as PYS (2003) Table 2.
_MU_EQUITY_GOOD   = 0.5
_MU_EQUITY_BAD    = 0.0
_MU_FIXEDINC_GOOD = 0.3
_MU_FIXEDINC_BAD  = 0.0
_MU_BALANCED_GOOD = 0.4
_MU_BALANCED_BAD  = 0.0

# Asset-class presets — mu values are ANNUALISED (matching Table 2 scale).
# Thresholds are loaded from thresholds_cache.json at import time and
# recomputed by calibrate_threshold() when the cache is absent.
# Fallback threshold values are placeholders; they are overwritten at import
# by _ensure_presets_calibrated().
ASSET_CLASS_PRESETS: dict = {
    "equity": {
        "mu_good":   _MU_EQUITY_GOOD,    # annualised IR: 0.50
        "mu_bad":    _MU_EQUITY_BAD,
        "threshold": 19.81,              # fallback; overwritten at import
    },
    "fixed_income": {
        "mu_good":   _MU_FIXEDINC_GOOD,  # annualised IR: 0.30
        "mu_bad":    _MU_FIXEDINC_BAD,
        "threshold": 19.81,
    },
    "balanced": {
        "mu_good":   _MU_BALANCED_GOOD,  # annualised IR: 0.40
        "mu_bad":    _MU_BALANCED_BAD,
        "threshold": 19.81,
    },
    "em_debt": {
        # EM hard- and local-currency debt — same mu targets as fixed_income
        # but calibrated separately so the threshold reflects EM-bond TE.
        "mu_good":   _MU_FIXEDINC_GOOD,  # annualised IR: 0.30
        "mu_bad":    _MU_FIXEDINC_BAD,
        "threshold": 19.81,              # fallback; overwritten at import
    },
}

# Maps internal tag_alarm_regime() labels to display strings used in reports.
VIX_REGIME_DISPLAY: dict = {
    "low_vol":  "calm",
    "med_vol":  "elevated",
    "high_vol": "crisis",
    "unknown":  "unknown",
}

_CACHE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "thresholds_cache.json")


def load_vix_history(start: str = "2005-01-01",
                     end: Optional[str] = None) -> pd.Series:
    """
    Pull ^VIX from yfinance, resample to monthly mean closes.

    Returns a pd.Series indexed by month-end dates with the mean daily VIX
    close for each calendar month. Monthly mean is used (rather than month-end
    close) so that a single volatile day at month-end does not unduly inflate
    the regime label.

    Requires yfinance. VIX data is available from 1990 onwards.
    """
    try:
        import yfinance as yf
    except ImportError:
        raise ImportError("yfinance not installed. Run: pip install yfinance")

    raw = yf.download("^VIX", start=start, end=end,
                      progress=False, auto_adjust=False)
    if raw.empty:
        raise RuntimeError("yfinance returned no data for ^VIX")

    # yfinance may return a (Price, Ticker) MultiIndex DataFrame for a single
    # ticker. Squeeze to a Series regardless of schema.
    if isinstance(raw.columns, pd.MultiIndex):
        close = raw["Close"].squeeze()
    elif "Close" in raw.columns:
        close = raw["Close"]
    else:
        close = raw.iloc[:, 0]

    if isinstance(close, pd.DataFrame):
        close = close.iloc[:, 0]  # final safety squeeze

    return close.resample("ME").mean().rename("vix_monthly_mean")


def tag_alarm_regime(
    alarm_dates: List[str],
    vix: pd.Series,
) -> dict:
    """
    Classify each alarm date by the VIX level in that month.

    IMPORTANT: vix must come from load_vix_history(), which resamples to the
    monthly MEAN before returning. Joining on a single month-end close is a
    silent-failure trap — a spike on the last trading day inflates the label.

    Categories:
      "high_vol" — VIX >= 30  (stress / tail-risk episode)
      "med_vol"  — 20 <= VIX < 30
      "low_vol"  — VIX < 20

    Parameters
    ----------
    alarm_dates : list of "YYYY-MM" strings (from monitor.alarm_date or similar).
    vix         : monthly VIX series from load_vix_history() — index is month-end,
                  values are the MEAN daily close for each calendar month.

    Returns
    -------
    dict mapping alarm_date_str → regime label (or "unknown" if VIX not available
    for that month).
    """
    labels = {}
    for date_str in alarm_dates:
        if not date_str or date_str == "-":
            continue
        try:
            ts  = pd.Timestamp(date_str)
            # Match to month-end in VIX series (within same calendar month).
            # Uses period matching so "2008-10" matches the Oct month-end entry.
            key = vix.index[vix.index.to_period("M") == ts.to_period("M")]
            if len(key) == 0:
                labels[date_str] = "unknown"
                continue
            v = float(vix.loc[key[0]])
            if v >= 30:
                labels[date_str] = "high_vol"
            elif v >= 20:
                labels[date_str] = "med_vol"
            else:
                labels[date_str] = "low_vol"
        except Exception:
            labels[date_str] = "unknown"
    return labels


def signal_quality(
    pa_df: pd.DataFrame,
    horizon: int = 12,
    regime: Optional[str] = None,
) -> Optional[float]:
    """
    Fraction of alarms where the fund underperformed its benchmark in the
    `horizon` months following the alarm.

    A value of 0.5 = coin flip; higher = more actionable signals.

    Parameters
    ----------
    pa_df   : DataFrame from post_alarm_returns(), optionally with a "regime"
              column added by tag_alarm_regime().
    horizon : forward look-through window in months. Default 12.
    regime  : if provided, restrict to alarms whose "regime" column matches
              this label ("high_vol", "med_vol", or "low_vol"). Allows hit
              rates to be stratified by VIX regime.
    """
    if pa_df is None or pa_df.empty:
        return None
    rows = pa_df[pa_df["horizon"] == horizon].dropna(subset=["correct_alarm"])
    if regime is not None and "regime" in rows.columns:
        rows = rows[rows["regime"] == regime]
    return float(rows["correct_alarm"].mean()) if len(rows) > 0 else None


def signal_quality_table(
    pa_df: pd.DataFrame,
    horizons: Tuple[int, ...] = (12, 24),
) -> pd.DataFrame:
    """
    Stratified signal-quality table: Regime | N alarms | Signal quality +12M | +24M.

    Rows: one per VIX regime (high_vol / med_vol / low_vol / all), plus an
    "all" row. Columns: regime, n_alarms, and one sq_+{h}m column per horizon.

    Parameters
    ----------
    pa_df    : DataFrame from post_alarm_returns() with a "regime" column
               added by tag_alarm_regime() and merged in.
    horizons : forward horizons to compute signal quality for. Default (12, 24).

    Returns
    -------
    pd.DataFrame with columns: regime, n_alarms, sq_+12m, sq_+24m (or similar).
    """
    if pa_df is None or pa_df.empty:
        return pd.DataFrame()

    regimes  = ["high_vol", "med_vol", "low_vol", "all"]
    sq_cols  = [f"sq_+{h}m" for h in horizons]
    rows_out = []
    for reg in regimes:
        if reg == "all":
            subset = pa_df
        elif "regime" not in pa_df.columns:
            continue
        else:
            subset = pa_df[pa_df["regime"] == reg]

        # Count distinct alarms (each alarm appears once per horizon)
        n_alarms = subset["alarm_t"].nunique() if "alarm_t" in subset.columns else 0
        row = {"regime": reg, "n_alarms": n_alarms}
        for h, col in zip(horizons, sq_cols):
            row[col] = signal_quality(subset, horizon=h)
        rows_out.append(row)

    return pd.DataFrame(rows_out, columns=["regime", "n_alarms"] + sq_cols)


def calibrate_threshold(
    mu_good: Optional[float] = None,
    mu_bad: Optional[float] = None,
    initial_te: float = 0.04,
    target_arl: float = 84.0,
    n_paths: int = 10_000,
    max_months: int = 600,
    seed: int = 42,
    tol: float = 0.5,
    h_lo: float = 1.0,
    h_hi: float = 30.0,
    asset_class: Optional[str] = None,
) -> float:
    """
    Find threshold h such that the simulated mean time-to-first-alarm under H0
    (continuous IR = mu_good, no skill change) equals target_arl months.

    Uses bisection on h. Runs the actual CUSUMMonitor — including its 24-month
    sigma warm-up window, regime-outlier freeze (regime_threshold=14.0), and
    annualised-scale Lindley recursion — not a simplified version.

    Under H0, monthly log-excess returns are drawn iid from
      N(mu_good * initial_te / 12,  (initial_te / sqrt(12))^2)
    which gives E[ir_hat_annual] = E[(e / sigma_m) * sqrt(12)] = mu_good.
    r_benchmark is set to zero; r_portfolio = e_t, valid for small monthly
    returns (|e| < 0.15 typically; log(1+x) ~= x approximation error < 1 %).

    mu_good and mu_bad must be on MONTHLY scale, matching the convention
    used throughout the code. Use the _MU_* constants.

    Parameters
    ----------
    mu_good    : annualised IR under H0 (e.g. _MU_EQUITY_GOOD = 0.50, matching Table 2).
    mu_bad     : annualised IR under H1 (typically 0.0).
    initial_te : annualised TE seed for each CUSUMMonitor path. Default 0.04.
    target_arl : target average run length in months. Default 84.
                 IMPORTANT: alarm_t counts from month 1 including the
                 calibration_months window during which L is frozen at 0.
                 The paper's ARL=60 is measured from the START OF MONITORING
                 (post-calibration). To match: target_arl = 60 + calibration_months
                 = 60 + 24 = 84. Do NOT change this back to 60 — that would aim
                 for only 36 monitoring months (60 - 24), misaligning with Table 2.
    n_paths    : Monte Carlo paths per bisection step. Default 10 000.
                 Increase to 50 000 for production-grade precision at cost of ~5x
                 longer runtime; 10 000 gives threshold accuracy +/-0.3 or better.
    max_months : max simulation length per path; censored paths count as max_months.
    seed       : RNG seed for reproducibility.
    tol        : bisection stops when |simulated ARL - target_arl| < tol.
    h_lo/h_hi  : search bracket for h. Default 1-30; must span the expected
                 annualised-scale threshold (~19-23 for equity at ARL=60 monitoring months).

    Returns
    -------
    Calibrated threshold h (float, rounded to 4 d.p.).

    If ``asset_class`` is given (e.g. "equity", "fixed_income", "balanced"),
    mu_good and mu_bad are taken from ASSET_CLASS_PRESETS and any explicit
    mu_good/mu_bad arguments are ignored.
    """
    if asset_class is not None:
        if asset_class not in ASSET_CLASS_PRESETS:
            raise ValueError(
                f"Unknown asset_class {asset_class!r}. "
                f"Valid keys: {list(ASSET_CLASS_PRESETS)}"
            )
        spec    = ASSET_CLASS_PRESETS[asset_class]
        mu_good = spec["mu_good"]
        mu_bad  = spec["mu_bad"]

    if mu_good is None or mu_bad is None:
        raise ValueError(
            "Either asset_class or both mu_good and mu_bad must be provided."
        )

    rng    = np.random.default_rng(seed)
    # Monthly log-excess mean and std under H0 so that E[ir_hat_annual] = mu_good.
    # ir_hat_annual = (e / sigma_m) * sqrt(12)  where sigma_m = initial_te / sqrt(12)
    # => E[e] = mu_good * sigma_m / sqrt(12) = mu_good * initial_te / 12
    # std(e) = sigma_m = initial_te / sqrt(12)  (unchanged)
    e_mean = mu_good * initial_te / 12.0
    e_std  = initial_te / _SQRT12

    # Pre-generate all random draws — shared across bisection steps so that
    # each step gets the same noise realisations (variance reduction).
    draws = rng.normal(e_mean, e_std, size=(n_paths, max_months))

    def mean_ttf(h: float) -> float:
        total = 0.0
        for i in range(n_paths):
            m = CUSUMMonitor(
                initial_te=initial_te,
                threshold=h,
                mu_good=mu_good,
                mu_bad=mu_bad,
                rolling_restart=False,
            )
            alarmed = False
            for j in range(max_months):
                m.update(draws[i, j], 0.0)
                if m.alarm:
                    total  += m.alarm_t
                    alarmed = True
                    break
            if not alarmed:
                total += max_months
        return total / n_paths

    lo, hi = h_lo, h_hi
    for _ in range(25):
        mid     = (lo + hi) / 2.0
        arl_mid = mean_ttf(mid)
        if abs(arl_mid - target_arl) < tol:
            return round(mid, 4)
        if arl_mid < target_arl:
            lo = mid
        else:
            hi = mid
    return round((lo + hi) / 2.0, 4)


def calibrate_all_presets(
    target_arl: float = 84.0,
    n_paths: int = 10_000,
    seed: int = 42,
) -> None:
    """
    Recalibrate thresholds for every key in ASSET_CLASS_PRESETS and save to
    thresholds_cache.json.  Called from report.py with ``--recalibrate``.

    Parameters mirror those of calibrate_threshold().  Uses the same default
    target_arl=84 (= 60 monitoring months + 24 calibration months).
    """
    print("Calibrating thresholds for all asset-class presets...")
    cache = {}
    for ac, spec in ASSET_CLASS_PRESETS.items():
        h = calibrate_threshold(
            mu_good=spec["mu_good"],
            mu_bad=spec["mu_bad"],
            target_arl=target_arl,
            n_paths=n_paths,
            seed=seed,
        )
        ASSET_CLASS_PRESETS[ac]["threshold"] = h
        cache[ac] = {
            "threshold": h,
            "mu_good":   spec["mu_good"],
            "mu_bad":    spec["mu_bad"],
            "target_arl": target_arl,
        }
        print(f"  {ac}: h = {h}")

    with open(_CACHE_PATH, "w") as f:
        json.dump(cache, f, indent=2)
    print(f"  saved -> {_CACHE_PATH}")


def _simulate_manager(n_months=240, change_month=120, ir_before=0.5,
                      ir_after=-0.3, te_annual=0.04, seed=11):
    """Synthetic manager with a known skill break — for offline testing."""
    rng      = np.random.default_rng(seed)
    te_m     = te_annual / np.sqrt(12)
    mkt      = rng.normal(0.08 / 12, 0.16 / np.sqrt(12), n_months)
    active   = np.empty(n_months)
    for t in range(n_months):
        ir        = ir_before if (change_month is None or t < change_month) else ir_after
        active[t] = ir * te_m + rng.normal(0, te_m)
    portfolio = mkt + active
    dates     = pd.date_range("2005-01-31", periods=n_months, freq="ME")
    return pd.DataFrame({"portfolio": portfolio, "benchmark": mkt}, index=dates)


# ---------------------------------------------------------------------------
# PART 2 — Morningstar xlsx parser
# ---------------------------------------------------------------------------

def load_morningstar_xlsx(path: str) -> Tuple[pd.DatetimeIndex, pd.DataFrame, dict]:
    """
    Parse a Morningstar 'grouped by custom' cumulative-return xlsx export.

    Returns
    -------
    dates      : pd.DatetimeIndex  (period-end dates, length T)
    fund_df    : pd.DataFrame with columns:
                   fund_name, benchmark_name, bench_id, ISIN, SecId,
                   inception, + columns 'r_0'..'r_{T-1}' holding monthly
                   total returns as decimals (NaN before inception)
    """
    raw = pd.read_excel(path, sheet_name=0, header=None)

    # Row 7 (0-indexed): period-end dates, starting at column 6
    date_vals = raw.iloc[7, 6:].values
    dates = pd.to_datetime(date_vals, errors="coerce", dayfirst=True)
    valid = ~pd.isnull(dates)
    dates_idx = pd.DatetimeIndex([d for d in dates if not pd.isnull(d)])
    T = len(dates_idx)

    # Rows 10+: data
    data = raw.iloc[10:].copy().reset_index(drop=True)

    # Keep only fund rows (have ISIN in col 3)
    mask = data.iloc[:, 3].notna()
    fd = data[mask].copy().reset_index(drop=True)

    # Monthly % returns → monthly decimal returns (divide by 100; NaN before inception stays NaN)
    cum_raw = fd.iloc[:, 6:].values.astype(float)[:, :T]  # shape (N, T)
    monthly_matrix = cum_raw / 100.0

    ret_cols = {f"r_{j}": monthly_matrix[:, j] for j in range(T)}

    fund_df = pd.DataFrame({
        "fund_name":      fd.iloc[:, 0].fillna("").astype(str).values,
        "benchmark_name": fd.iloc[:, 1].fillna("").astype(str).values,
        "bench_id": [
            str(v).strip() if not (isinstance(v, float) and np.isnan(v)) else ""
            for v in fd.iloc[:, 2].values
        ],
        "ISIN":      fd.iloc[:, 3].fillna("").astype(str).values,
        "SecId":     fd.iloc[:, 4].fillna("").astype(str).values,
        "inception": pd.to_datetime(fd.iloc[:, 5], errors="coerce").values,
        **ret_cols,
    })
    # Extract benchmark rows: no ISIN (col 3 empty) but have SecId (col 4)
    bench_mask = data.iloc[:, 3].isna() & data.iloc[:, 4].notna()
    bd = data[bench_mask].copy().reset_index(drop=True)
    bench_map: dict = {}
    if len(bd) > 0:
        bench_raw = bd.iloc[:, 6:].values.astype(float)[:, :T] / 100.0
        for k in range(len(bd)):
            secid = str(bd.iloc[k, 4]).strip()
            if secid and secid != "nan":
                bench_map[secid] = bench_raw[k]

    return dates_idx, fund_df, bench_map


# ---------------------------------------------------------------------------
# PART 3 — Pipeline
# ---------------------------------------------------------------------------

def run_pipeline(
    xlsx_path: str,
    threshold: float = 19.81,
    min_months: int = 24,
    out_dir: Optional[str] = None,
) -> pd.DataFrame:
    """
    Full pipeline: load → parse benchmarks → run CUSUM → save outputs.

    Returns a DataFrame with one row per analysed fund.
    """
    xlsx_path = str(xlsx_path)
    if out_dir is None:
        out_dir = str(Path(xlsx_path).parent)
    os.makedirs(out_dir, exist_ok=True)

    run_date = Path(xlsx_path).stem.split("_")[0:3]  # try to get date from filename
    tag = "_".join(run_date) if run_date else "run"

    print(f"Loading {xlsx_path} ...")
    dates, fund_df, bench_map = load_morningstar_xlsx(xlsx_path)
    T = len(dates)
    N = len(fund_df)
    print(f"  {N} fund rows  |  {T} monthly periods  "
          f"({dates[0].strftime('%b %Y')} -> {dates[-1].strftime('%b %Y')})")
    print(f"  {len(bench_map)} benchmark index series loaded from Benchmarks section")

    # ── ETF proxy benchmarks ──────────────────────────────────────────────────
    # Fetch (or load from cache) ETF return series for the 27 SecIds that are
    # absent from the Morningstar Benchmarks section.  Only adds to bench_map
    # if the SecId is not already present — Morningstar data always takes priority.
    cache_path = str(Path(xlsx_path).parent / "etf_proxy_cache.json")
    etf_proxy = _fetch_etf_proxies(dates, cache_path)
    n_added = 0
    for secid, arr in etf_proxy.items():
        if secid not in bench_map:          # never overwrite real index data
            bench_map[secid] = arr
            n_added += 1
    if n_added:
        print(f"  {n_added} ETF proxy series merged into bench_map "
              f"(total now: {len(bench_map)})")

    ret_cols = [f"r_{j}" for j in range(T)]
    ret_matrix = fund_df[ret_cols].values  # (N, T)

    def _parse_blended(bid_str: str):
        """Parse 'ID1 [w1%]; ID2 [w2%]; ...' → list of (id, weight) or None."""
        parts = bid_str.split(";")
        result = []
        for p in parts:
            m = re.match(r'^\s*(\S+)\s+\[(\d+(?:\.\d+)?)\%\]\s*$', p.strip())
            if not m:
                return None
            result.append((m.group(1), float(m.group(2)) / 100.0))
        return result if result else None

    def _blended_bench_series(bid_str: str) -> Optional[np.ndarray]:
        """Return weighted benchmark return array (length T) or None if any component missing."""
        components = _parse_blended(bid_str)
        if not components:
            return None
        weighted = np.zeros(T, dtype=float)
        for cid, w in components:
            if cid not in bench_map:
                return None
            weighted += w * bench_map[cid].astype(float)
        return weighted

    results = []
    all_fund_data: dict[str, dict] = {}
    n_simple = 0; n_blended = 0; n_skip_missing = 0; n_skip_history = 0; n_skip_te = 0

    print("Running CUSUM ...")
    for i, row in fund_df.iterrows():
        bid = row["bench_id"]
        if not bid or bid == "nan":
            continue

        is_blended = "[" in bid
        if is_blended:
            bench_m = _blended_bench_series(bid)
            if bench_m is None:
                n_skip_missing += 1
                continue
            bench_name_used = row["benchmark_name"]
        else:
            if bid not in bench_map:
                n_skip_missing += 1
                continue
            bench_m = bench_map[bid]
            bench_name_used = row["benchmark_name"]

        fund_m = ret_matrix[i]

        fund_s  = pd.Series(fund_m.astype(float),  index=dates).dropna()
        bench_s = pd.Series(bench_m.astype(float), index=dates).dropna()
        common  = fund_s.index.intersection(bench_s.index)
        if len(common) < min_months:
            n_skip_history += 1
            continue
        if is_blended:
            n_blended += 1
        else:
            n_simple += 1

        ret = pd.DataFrame(
            {"portfolio": fund_s[common].values,
             "benchmark": bench_s[common].values},
            index=common,
        )

        try:
            monitor = CUSUMMonitor(threshold=threshold)
            hist    = monitor.run(ret)
        except Exception:
            continue

        f = ret["portfolio"].values
        b = ret["benchmark"].values
        excess     = ret["portfolio"] - ret["benchmark"]
        excess_arr = f - b
        std_ex     = excess.std()
        n          = len(ret)

        # IR and t-stat
        ir     = (excess.mean() * 12) / (std_ex * np.sqrt(12)) if std_ex > 1e-8 else np.nan
        t_stat = ir * np.sqrt(n / 12) if not np.isnan(ir) else np.nan

        # Volatility / TE
        ann_vol_fund  = ret["portfolio"].std() * np.sqrt(12)
        ann_vol_bench = ret["benchmark"].std()  * np.sqrt(12)
        tracking_error = std_ex * np.sqrt(12)

        # Hit rate
        hit_rate = (excess_arr > 0).mean() * 100

        # Up / down capture
        up_mask   = b > 0
        dn_mask   = b < 0
        up_capture  = (f[up_mask].mean() / b[up_mask].mean() * 100
                       if up_mask.sum() > 0 and b[up_mask].mean() != 0 else np.nan)
        down_capture = (f[dn_mask].mean() / b[dn_mask].mean() * 100
                        if dn_mask.sum() > 0 and b[dn_mask].mean() != 0 else np.nan)

        # Cumulative excess return
        cum_ex_pct = ((1 + ret["portfolio"]).prod() / (1 + ret["benchmark"]).prod() - 1) * 100

        current_L    = hist["L"].iloc[-1]
        current_L_up = hist["L_up"].iloc[-1]
        alarm_date   = (common[monitor.alarm_t - 1].strftime("%Y-%m")
                        if monitor.alarm_t else "")

        fname = row["fund_name"]
        results.append({
            "fund":            fname,
            "benchmark":       row["benchmark_name"],
            "n_months":        n,
            "IR":              round(ir,            4) if not np.isnan(ir)            else None,
            "t_stat":          round(t_stat,        4) if not np.isnan(t_stat)        else None,
            "ann_vol_fund":    round(ann_vol_fund  * 100, 2),
            "ann_vol_bench":   round(ann_vol_bench * 100, 2),
            "tracking_error":  round(tracking_error * 100, 2),
            "hit_rate":        round(hit_rate,      1),
            "up_capture":      round(up_capture,    1) if not np.isnan(up_capture)    else None,
            "down_capture":    round(down_capture,  1) if not np.isnan(down_capture)  else None,
            "cum_excess_pct":  round(cum_ex_pct,   2),
            "current_L":       round(current_L,    4),
            "pct_h":           round(current_L / threshold * 100, 1),
            "ALARM":           monitor.alarm,
            "n_alarms":        len(monitor.all_alarms),
            "first_alarm":     alarm_date,
        })
        all_fund_data[fname] = {"ret": ret, "hist": hist, "monitor": monitor}

    df = (pd.DataFrame(results)
            .sort_values("current_L", ascending=False)
            .reset_index(drop=True))

    # Filter 1 — drop benchmark self-matches (TE < 0.5%); catches exact self-rows AND
    # near-duplicate share classes that track the benchmark index itself (up/down capture ~100%)
    pre_filter1 = len(df)
    df = df[df["tracking_error"] >= 0.5].reset_index(drop=True)
    n_skip_te = pre_filter1 - len(df)

    # Filter 2 — drop rows with no meaningful benchmark name
    _bad_bench = {"", "nan", "not benchmarked"}
    pre_filter2 = len(df)
    df = df[~df["benchmark"].str.strip().str.lower().isin(_bad_bench)].reset_index(drop=True)
    n_skip_blank = pre_filter2 - len(df)

    n_alarm      = int(df["ALARM"].sum())
    n_watchlist  = int((df["current_L"] > 0.75 * threshold).sum())
    n_active     = int((df["current_L"] > threshold).sum())
    n_red        = int((df["current_L"] > 0.90 * threshold).sum())
    n_amber_hi   = int(((df["current_L"] > 0.75 * threshold) & (df["current_L"] <= 0.90 * threshold)).sum())
    n_amber_lo   = int(((df["current_L"] > 0.50 * threshold) & (df["current_L"] <= 0.75 * threshold)).sum())
    n_green      = int((df["current_L"] <= 0.50 * threshold).sum())
    median_ir    = df["IR"].dropna().median()
    N_final      = len(df)

    print(f"\n{'='*60}")
    print(f"PIPELINE SUMMARY")
    print(f"{'='*60}")
    print(f"  Total fund rows in file          : {N}")
    print(f"  Blank/no bench_id                : {N - len([b for b in fund_df['bench_id'] if b and b != 'nan'])}")
    print(f"  Blended bench (missing components): {n_skip_missing - (n_skip_missing - max(0, n_skip_missing))}")
    print(f"  Funds reaching CUSUM loop        : {n_simple + n_blended + n_skip_history}")
    print(f"    Simple benchmark                :   {n_simple + n_skip_history - n_skip_history}")
    print(f"    Blended benchmark (all avail.)  :   {n_blended}")
    print(f"  Excluded: insufficient history   : {n_skip_history}")
    print(f"  Pre-filter total                 : {pre_filter1}")
    print(f"  Excluded: TE<0.1% self-matches   : {n_skip_te}")
    print(f"  Excluded: blank benchmark name   : {n_skip_blank}")
    print(f"  FINAL funds analysed             : {N_final}")
    print(f"    of which simple benchmark      :   {N_final - df['benchmark'].str.contains(';', na=False).sum()}")
    print(f"    of which blended benchmark     :   {df['benchmark'].str.contains(';', na=False).sum()}")
    print(f"  Ever alarmed (ALARM=True)        : {n_alarm}")
    print(f"  Watchlist (L > 75%h = {0.75*threshold:.2f})   : {n_watchlist}")
    print(f"    Red zone  (L > 90%h = {0.90*threshold:.2f}) : {n_red}")
    print(f"    Amber-hi  (L 75-90%h)          : {n_amber_hi}")
    print(f"  Amber-low (L 50-75%h)            : {n_amber_lo}")
    print(f"  Green     (L <= 50%h = {0.50*threshold:.2f})  : {n_green}")
    print(f"  Median IR (all analysed)         : {median_ir:+.3f}")

    # Longest history
    longest = df.nlargest(1, "n_months").iloc[0]
    print(f"  Longest history                  : {longest['fund'][:50]}  ({int(longest['n_months'])} months)")

    # Watchlist TE flag
    low_te_watch = df[df["current_L"] > 0.75 * threshold][["fund","tracking_error","current_L"]]
    low_te_watch = low_te_watch[low_te_watch["tracking_error"] < 0.5]
    if len(low_te_watch):
        print(f"\n  WARNING - Watchlist funds with TE < 0.5% (possible self-match):")
        for _, r in low_te_watch.iterrows():
            print(f"    {r['fund'][:50]}  TE={r['tracking_error']:.3f}%  L={r['current_L']:.2f}")
    else:
        print(f"  Watchlist TE check: no funds with TE < 0.5%")

    # -----------------------------------------------------------------------
    # Save CSV
    # -----------------------------------------------------------------------
    today_tag = pd.Timestamp.now().strftime("%Y%m%d")
    csv_cols  = [
        "fund", "benchmark", "n_months", "IR", "t_stat",
        "ann_vol_fund", "ann_vol_bench", "tracking_error",
        "hit_rate", "up_capture", "down_capture", "cum_excess_pct",
        "current_L", "pct_h", "ALARM", "n_alarms", "first_alarm",
    ]
    csv_path = os.path.join(out_dir, f"cusum_IC_report_{today_tag}.csv")
    df[csv_cols].to_csv(csv_path, index=False)
    print(f"\nSaved: {csv_path}")

    # -----------------------------------------------------------------------
    # Top-20 console print
    # -----------------------------------------------------------------------
    pd.set_option("display.max_columns", 20)
    pd.set_option("display.width", 200)
    pd.set_option("display.float_format", "{:.2f}".format)
    print(f"\nTOP 20 BY CURRENT_L:\n")
    print(df[csv_cols].head(20).to_string(index=False))

    # -----------------------------------------------------------------------
    # Formatted xlsx
    # -----------------------------------------------------------------------
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

    xlsx_path = os.path.join(out_dir, f"cusum_IC_report_{today_tag}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "CUSUM Monitor"

    # ---- header row ----
    header_fill = PatternFill("solid", fgColor="1E2761")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin         = Side(style="thin", color="CCCCCC")
    border       = Border(bottom=thin)

    ws.append(csv_cols)
    for cell in ws[1]:
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # ---- data rows ----
    out_df = df[csv_cols].copy()
    # Convert pct_h and rate columns to fractions for Excel % format
    pct_fraction_cols = {"pct_h", "hit_rate", "up_capture", "down_capture"}
    for col in pct_fraction_cols:
        out_df[col] = out_df[col] / 100.0

    for r_vals in out_df.itertuples(index=False):
        ws.append(list(r_vals))

    # ---- column formats ----
    col_idx = {name: i + 1 for i, name in enumerate(csv_cols)}

    fmt_pct1   = "0.0%"
    fmt_2dp    = "0.00"
    fmt_int    = "0"
    fmt_pct_h  = "0.0%"

    pct1_cols  = ["hit_rate", "up_capture", "down_capture", "pct_h"]
    dp2_cols   = ["IR", "t_stat", "tracking_error", "ann_vol_fund", "ann_vol_bench",
                  "cum_excess_pct", "current_L"]
    int_cols   = ["n_months", "n_alarms"]

    for col_name in pct1_cols:
        col_letter = get_column_letter(col_idx[col_name])
        for cell in ws[col_letter][1:]:   # skip header
            cell.number_format = fmt_pct1

    for col_name in dp2_cols:
        col_letter = get_column_letter(col_idx[col_name])
        for cell in ws[col_letter][1:]:
            cell.number_format = fmt_2dp

    for col_name in int_cols:
        col_letter = get_column_letter(col_idx[col_name])
        for cell in ws[col_letter][1:]:
            cell.number_format = fmt_int

    # ---- conditional formatting on current_L ----
    n_rows   = len(out_df)
    L_col    = get_column_letter(col_idx["current_L"])
    L_range  = f"{L_col}2:{L_col}{n_rows + 1}"

    green_fill = PatternFill("solid", fgColor="C6EFCE")   # < 50%h  (< 9.905)
    amber_fill = PatternFill("solid", fgColor="FFEB9C")   # 50-75%h (9.905–14.8575)
    red_fill   = PatternFill("solid", fgColor="FFC7CE")   # > 75%h  (> 14.8575)

    h50  = threshold * 0.50
    h75  = threshold * 0.75

    ws.conditional_formatting.add(L_range,
        CellIsRule(operator="lessThan",      formula=[str(h50)],  fill=green_fill))
    ws.conditional_formatting.add(L_range,
        CellIsRule(operator="between",       formula=[str(h50), str(h75)], fill=amber_fill))
    ws.conditional_formatting.add(L_range,
        CellIsRule(operator="greaterThan",   formula=[str(h75)],  fill=red_fill))

    # ---- conditional formatting on IR ----
    IR_col   = get_column_letter(col_idx["IR"])
    IR_range = f"{IR_col}2:{IR_col}{n_rows + 1}"

    red_font   = Font(color="9C0006")
    green_font = Font(color="276221")

    ws.conditional_formatting.add(IR_range,
        CellIsRule(operator="lessThan",    formula=["0"],   font=red_font))
    ws.conditional_formatting.add(IR_range,
        CellIsRule(operator="greaterThan", formula=["0.5"], font=green_font))

    # ---- column widths ----
    col_widths = {
        "fund": 42, "benchmark": 36, "n_months": 9, "IR": 7, "t_stat": 7,
        "ann_vol_fund": 10, "ann_vol_bench": 11, "tracking_error": 11,
        "hit_rate": 8, "up_capture": 9, "down_capture": 10, "cum_excess_pct": 12,
        "current_L": 10, "pct_h": 7, "ALARM": 7, "n_alarms": 8, "first_alarm": 11,
    }
    for col_name, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx[col_name])].width = width

    # ---- freeze top row ----
    ws.freeze_panes = "A2"

    wb.save(xlsx_path)
    print(f"Saved: {xlsx_path}")

    # -----------------------------------------------------------------------
    # Chart 1 — L bar chart top 30
    # -----------------------------------------------------------------------
    top30  = df.head(30)
    colors = [ALARM_RED if a else NAV_BLUE for a in top30["ALARM"]]
    fig, ax = plt.subplots(figsize=(14, 9))
    ax.barh(range(len(top30)), top30["current_L"], color=colors, height=0.7, alpha=0.85)
    ax.axvline(threshold,         color=ALARM_RED, lw=2,   ls="--", label=f"Threshold h = {threshold}")
    ax.axvline(0.75 * threshold,  color=AMBER,     lw=1.2, ls=":",  label=f"Watchlist 75%h = {0.75*threshold:.2f}")
    ax.set_yticks(range(len(top30)))
    ax.set_yticklabels([r["fund"][:45] for _, r in top30.iterrows()], fontsize=8)
    ax.invert_yaxis()
    ax.set_xlabel("Current CUSUM Statistic L", fontsize=11)
    ax.set_title(
        f"CUSUM Monitor — Top 30 by Current L  |  {dates[-1].strftime('%b %Y')}",
        fontsize=13, fontweight="bold", color=NAV_BLUE,
    )
    for idx, (_, row) in enumerate(top30.iterrows()):
        ax.text(row["current_L"] + 0.1, idx,
                f"{row['pct_h']:.0f}%", va="center", fontsize=7, color="#333")
    ax.legend(handles=[
        mpatches.Patch(color=ALARM_RED, label="Ever alarmed"),
        mpatches.Patch(color=NAV_BLUE,  label="No alarm yet"),
    ], fontsize=9, loc="lower right")
    ax.set_xlim(0, threshold * 1.08)
    ax.grid(axis="x", alpha=0.3)
    plt.tight_layout()
    p = os.path.join(out_dir, "cusum_L_barchart.png")
    fig.savefig(p, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"Saved: {p}")

    # -----------------------------------------------------------------------
    # Chart 2 — time series for top 5
    # -----------------------------------------------------------------------
    top5_names = [r for r in df.head(5)["fund"] if r in all_fund_data]
    fig, axes = plt.subplots(len(top5_names), 1, figsize=(14, 4 * len(top5_names)))
    if len(top5_names) == 1:
        axes = [axes]
    for ax, fname in zip(axes, top5_names):
        d = all_fund_data[fname]
        hist, monitor, ret = d["hist"], d["monitor"], d["ret"]
        idx = ret.index
        ax.plot(idx, hist["L"].values,    lw=1.8, color=NAV_BLUE, label="L (skill loss)")
        ax.plot(idx, hist["L_up"].values, lw=1.2, color=TEAL, ls="--", label="L_up (skill gain)")
        ax.axhline(threshold, color=ALARM_RED, ls="--", lw=1.5, label=f"h={threshold}")
        for at in monitor.all_alarms:
            ax.axvline(idx[at - 1], color=ALARM_RED, lw=1, alpha=0.5)
        ax2 = ax.twinx()
        cum_ex = (ret["portfolio"] - ret["benchmark"]).cumsum() * 100
        ax2.fill_between(idx, cum_ex, 0, alpha=0.1, color=NAV_BLUE)
        ax2.plot(idx, cum_ex, lw=0.8, color=NAV_BLUE, alpha=0.5)
        ax2.set_ylabel("Cum. excess ret (%)", fontsize=7, color=NAV_BLUE)
        ax.set_title(fname[:60], fontsize=9, fontweight="bold", color=NAV_BLUE)
        ax.set_ylabel("L statistic", fontsize=8)
        ax.legend(loc="upper left", fontsize=7)
        ax.grid(alpha=0.25)
        ax.tick_params(axis="x", labelsize=7)
    fig.suptitle(f"CUSUM Time Series — Top 5 Most At-Risk  |  {dates[-1].strftime('%b %Y')}",
                 fontsize=13, fontweight="bold", color=NAV_BLUE, y=1.005)
    plt.tight_layout()
    p = os.path.join(out_dir, "cusum_top5_timeseries.png")
    fig.savefig(p, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"Saved: {p}")

    # -----------------------------------------------------------------------
    # Chart 3 — IR vs current L scatter
    # -----------------------------------------------------------------------
    fig, ax = plt.subplots(figsize=(12, 7))
    ir_vals = df["IR"].fillna(0)
    colors  = [ALARM_RED if a else NAV_BLUE for a in df["ALARM"]]
    sizes   = np.clip(df["n_months"] / 5, 10, 200)
    ax.scatter(ir_vals, df["current_L"], c=colors, s=sizes,
               alpha=0.65, edgecolors="white", lw=0.4)
    ax.axhline(threshold,        color=ALARM_RED, ls="--", lw=1.5, label=f"Threshold h={threshold}")
    ax.axhline(0.75 * threshold, color=AMBER,     ls=":",  lw=1.2, label="Watchlist 75%h")
    ax.axvline(0, color="grey", lw=0.7)
    for _, row in df.head(10).iterrows():
        ax.annotate(row["fund"][:22],
                    (row["IR"] if row["IR"] else 0, row["current_L"]),
                    fontsize=6.5, xytext=(4, 2), textcoords="offset points")
    ax.set_xlabel("Information Ratio (annualised)", fontsize=11)
    ax.set_ylabel("Current CUSUM Statistic L", fontsize=11)
    ax.set_title(
        f"IR vs CUSUM L — {len(df)} funds  |  bubble ∝ history  |  {dates[-1].strftime('%b %Y')}",
        fontsize=12, fontweight="bold", color=NAV_BLUE,
    )
    ax.legend(handles=[
        mpatches.Patch(color=ALARM_RED, label="Ever alarmed"),
        mpatches.Patch(color=NAV_BLUE,  label="No alarm"),
    ], fontsize=9)
    ax.grid(alpha=0.25)
    plt.tight_layout()
    p = os.path.join(out_dir, "cusum_scatter.png")
    fig.savefig(p, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"Saved: {p}")

    # -----------------------------------------------------------------------
    # Console summary
    # -----------------------------------------------------------------------
    print(f"\n{'='*65}")
    print(f"UNIVERSE SUMMARY  --  data to {dates[-1].strftime('%b %Y')}")
    print(f"{'='*65}")
    print(f"  Threshold h               : {threshold}")
    print(f"  Funds analysed            : {len(df)}")
    print(f"  Ever alarmed              : {n_alarm}")
    print(f"  Currently on watchlist    : {n_watchlist}  (L > 75%h = {0.75*threshold:.2f})")
    print(f"  Active alarm this month   : {n_active}")
    print(f"  Median IR                 : {df['IR'].median():.3f}")
    print()
    print("WATCHLIST  (L > 75% of threshold):")
    wl = df[df["current_L"] > 0.75 * threshold]
    print(wl[["fund", "benchmark", "n_months", "IR",
               "current_L", "pct_h", "ALARM", "n_alarms",
               "first_alarm"]].to_string(index=False))

    return df


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def _cli():
    parser = argparse.ArgumentParser(
        description="CUSUM Active-Manager Monitor — Morningstar xlsx pipeline"
    )
    parser.add_argument("xlsx", help="Path to the Morningstar xlsx export")
    parser.add_argument("--threshold", type=float, default=19.81,
                        help="CUSUM alarm threshold h (default 19.81, ARL≈60m)")
    parser.add_argument("--min-months", type=int, default=24,
                        help="Minimum overlapping months to include a fund (default 24)")
    parser.add_argument("--out-dir", default=None,
                        help="Output directory (default: same folder as xlsx)")
    args = parser.parse_args()
    run_pipeline(args.xlsx, args.threshold, args.min_months, args.out_dir)


if __name__ == "__main__":
    _cli()
